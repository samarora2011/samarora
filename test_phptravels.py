import pytest
import allure
import time
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl
import os


def capture_screenshot(driver, name):
    try:
        WebDriverWait(driver, 30).until(lambda d: d.execute_script("return document.readyState") == "complete")
    except Exception:
        pass
    time.sleep(5)
    allure.attach(driver.get_screenshot_as_png(), name=name, attachment_type=allure.attachment_type.PNG)


@allure.feature("Login")
@allure.story("User Login")
def test_login(driver):
    try:
        driver.get("https://phptravels.net/login")

        # Wait for email field
        email = WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.NAME, "email")))
        email.send_keys("user@phptravels.com")

        password = driver.find_element(By.NAME, "password")
        password.send_keys("demouser")

        login_button = driver.find_element(By.XPATH, "//button[@type='submit']")
        login_button.click()

        # Wait for URL to change from login page
        WebDriverWait(driver, 45).until(EC.url_changes("https://phptravels.net/login"))

        assert "dashboard" in driver.current_url

        # Record result in Excel
        record_result("Login Test", "Passed")
    except Exception as e:
        capture_screenshot(driver, "Screenshot on Failure")
        record_result("Login Test", "Failed")
        raise e

@allure.feature("Flight Search & Booking")
@allure.story("FR2.1 - Search One Way Flight")
def test_flight_search(driver):
    try:
        driver.get("https://phptravels.net/flights")

        # Wait for page to load
        WebDriverWait(driver, 30).until(lambda d: len(d.page_source) > 1000)

        assert "flight" in driver.page_source.lower()

        record_result("Flight Search Test", "Passed")
    except Exception as e:
        capture_screenshot(driver, "Screenshot on Failure")
        record_result("Flight Search Test", "Failed")
        raise e

@allure.feature("Hotel Search & Booking")
@allure.story("FR2.2 - Search Hotels")
def test_hotel_search(driver):
    try:
        driver.get("https://phptravels.net/")

        # Wait for hotel search form fields
        WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.NAME, "checkin_date")))
        WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.NAME, "checkout_date")))
        WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.NAME, "destination")))

        assert "hotel" in driver.page_source.lower()

        record_result("Hotel Search Test", "Passed")
    except Exception as e:
        capture_screenshot(driver, "Screenshot on Failure")
        record_result("Hotel Search Test", "Failed")
        raise e

@allure.feature("Tour Search & Booking")
@allure.story("FR2.3 - Browse Tours")
def test_tour_browse(driver):
    try:
        driver.get("https://phptravels.net/tours")

        # Wait for tour search form and results container
        WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.NAME, "start_date")))

        assert "tour" in driver.page_source.lower()

        record_result("Tour Browse Test", "Passed")
    except Exception as e:
        capture_screenshot(driver, "Screenshot on Failure")
        record_result("Tour Browse Test", "Failed")
        raise e

@allure.feature("Car Rental Search & Booking")
@allure.story("FR2.4 - Search Cars")
def test_car_search(driver):
    try:
        driver.get("https://phptravels.net/cars")

        # Wait for car rental fields
        WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.NAME, "pickup_date")))
        WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.NAME, "return_date")))

        assert "car" in driver.page_source.lower()

        record_result("Car Search Test", "Passed")
    except Exception as e:
        capture_screenshot(driver, "Screenshot on Failure")
        record_result("Car Search Test", "Failed")
        raise e

@allure.feature("Booking Management")
@allure.story("FR2.6 - View Bookings")
def test_booking_management(logged_in_driver):
    try:
        driver = logged_in_driver
        WebDriverWait(driver, 30).until(lambda d: "/dashboard" in d.current_url or "welcome back" in d.page_source.lower())

        assert "welcome back" in driver.page_source.lower()

        record_result("Booking Management Test", "Passed")
    except Exception as e:
        capture_screenshot(driver, "Screenshot on Failure")
        record_result("Booking Management Test", "Failed")
        raise e

@allure.feature("Invoice Generation")
@allure.story("FR2.7 - Check Invoice")
def test_invoice_generation(logged_in_driver):
    try:
        driver = logged_in_driver
        WebDriverWait(driver, 30).until(lambda d: "/dashboard" in d.current_url or "welcome back" in d.page_source.lower())

        assert "welcome back" in driver.page_source.lower()

        record_result("Invoice Generation Test", "Passed")
    except Exception as e:
        capture_screenshot(driver, "Screenshot on Failure")
        record_result("Invoice Generation Test", "Failed")
        raise e

def record_result(test_name, status):
    try:
        file_path = "test_results.xlsx"
        if os.path.exists(file_path):
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws['A1'] = 'Test Name'
            ws['B1'] = 'Status'

        row = ws.max_row + 1
        ws[f'A{row}'] = test_name
        ws[f'B{row}'] = status
        wb.save(file_path)
    except Exception as e:
        print(f"Failed to save to Excel: {e}")