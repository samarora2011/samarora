from openpyxl import load_workbook

from utils.search_flights_data import SearchFlightsData
from utils.search_hotels_data import SearchHotelsData
from utils.search_tours_data import SearchToursData
from utils.search_transfers_data import SearchTransfersData


class XlsxReader:
    @staticmethod
    def get_xlsx_hotels_data():
        wb = load_workbook(f"./utils/search_hotels_input_data.xlsx")
        sheet = wb.active
        data = []

        for i in range(2, sheet.max_row + 1):
            search_hotels_data = SearchHotelsData(sheet.cell(i, 1).value,  # destination
                                                  sheet.cell(i, 2).value,  # check in date
                                                  sheet.cell(i, 3).value,  # check out date
                                                  int(sheet.cell(i, 4).value),  # adults number
                                                  int(sheet.cell(i, 5).value))  # kids number
            data.append(search_hotels_data)
        return data

    @staticmethod
    def get_xlsx_flights_data():
        wb = load_workbook(f"./utils/search_flights_input_data.xlsx")
        sheet = wb.active
        data = []

        for i in range(2, sheet.max_row + 1):
            search_flights_data = SearchFlightsData(sheet.cell(i, 1).value,  # cabin class
                                                    sheet.cell(i, 2).value,  # location from
                                                    sheet.cell(i, 3).value,  # location to
                                                    sheet.cell(i, 4).value,  # start year
                                                    sheet.cell(i, 5).value,  # start month
                                                    sheet.cell(i, 6).value,  # start day
                                                    sheet.cell(i, 7).value,  # end year
                                                    sheet.cell(i, 8).value,  # end month
                                                    sheet.cell(i, 9).value,  # end day
                                                    int(sheet.cell(i, 10).value),  # adults number
                                                    int(sheet.cell(i, 11).value),  # kids number
                                                    int(sheet.cell(i, 12).value))  # infants number
            data.append(search_flights_data)
        return data

    @staticmethod
    def get_xlsx_tours_data():
        wb = load_workbook(f"./utils/search_tours_input_data.xlsx")
        sheet = wb.active
        data = []

        for i in range(2, sheet.max_row + 1):
            search_tours_data = SearchToursData(sheet.cell(i, 1).value,  # destination
                                                sheet.cell(i, 2).value,  # tour type
                                                sheet.cell(i, 3).value,  # start year
                                                sheet.cell(i, 4).value,  # start month
                                                sheet.cell(i, 5).value,  # start day
                                                int(sheet.cell(i, 6).value))  # adults number
            data.append(search_tours_data)
        return data

    @staticmethod
    def get_xlsx_transfers_data():
        wb = load_workbook(f"./utils/search_transfers_input_data.xlsx")
        sheet = wb.active
        data = []

        for i in range(2, sheet.max_row + 1):
            search_transfers_data = SearchTransfersData(sheet.cell(i, 1).value,  # pick up location
                                                        sheet.cell(i, 2).value,  # drop off location
                                                        sheet.cell(i, 3).value,  # depart year
                                                        sheet.cell(i, 4).value,  # depart month
                                                        sheet.cell(i, 5).value,  # depart day
                                                        sheet.cell(i, 6).value,  # depart time
                                                        sheet.cell(i, 7).value,  # return year
                                                        sheet.cell(i, 8).value,  # return month
                                                        sheet.cell(i, 9).value,  # return day
                                                        sheet.cell(i, 10).value)  # return time
            data.append(search_transfers_data)
        return data
